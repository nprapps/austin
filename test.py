#!/usr/bin/env python

from openpyxl.reader.excel import load_workbook

from copytext import Copy

print 'BAD'

book = load_workbook('copy_bad.xlsx', data_only=True)

for sheet in book:
    print sheet.title
    print sheet.rows[1][0].internal_value

copy = Copy('copy_bad.xlsx')

print copy['content']
print unicode(copy['content']['brand'])

print ''
print 'GOOD'

book = load_workbook('copy_good.xlsx', data_only=True)

for sheet in book:
    print sheet.title
    print sheet.rows[1][0].internal_value

copy = Copy('copy_good.xlsx')

print copy['content']
print unicode(copy['content']['brand'])
